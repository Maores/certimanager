import openpyxl
import sys
import io
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def read_sheet_data(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows

def find_header_and_id_col(rows):
    """Find header row and ID column - broad search."""
    id_keywords = ['ת.ז', 'ת"ז', 'תז', 'מס זהות', 'מספר זהות', 'תעודת זהות', 'זהות/דרכון', 'מספר זהות/דרכון']
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            for kw in id_keywords:
                if kw in cell_str:
                    return i, j
    return None, None

def extract_employees(rows, header_idx, id_col_idx):
    """Extract unique IDs and employee count from rows after header."""
    ids = set()
    employee_count = 0
    data_rows = rows[header_idx + 1:]
    for row in data_rows:
        if all(c is None for c in row):
            continue
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty < 2:
            continue
        employee_count += 1
        if id_col_idx is not None and id_col_idx < len(row) and row[id_col_idx] is not None:
            id_val = str(row[id_col_idx]).strip()
            if id_val and id_val not in ('None', 'nan', '', '-'):
                if id_val.endswith('.0'):
                    id_val = id_val[:-2]
                ids.add(id_val)
    return employee_count, ids

def analyze_pikoh():
    path = r"C:\Users\maor4\OneDrive\Desktop\עותק של מאושרי_נתע_לשיבוץ_מעודכן.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results = {}
    all_ids = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = read_sheet_data(ws)
        header_idx, id_col_idx = find_header_and_id_col(rows)

        if header_idx is not None:
            employee_count, ids = extract_employees(rows, header_idx, id_col_idx)
            results[sheet_name] = {'count': employee_count, 'ids': ids}
            all_ids.update(ids)
        else:
            results[sheet_name] = {'count': 0, 'ids': set()}

    wb.close()
    return results, all_ids

def analyze_ka():
    path = r"C:\Users\maor4\OneDrive\Desktop\כ״א +משימות.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results = {}
    all_ids = set()

    # Primary cert type sheets (exclude summary, task aggregation, and task-by-owner sheets)
    cert_sheets = ['מאושרי נת״ע', 'מאושרי כביש 6', 'מאושרי כביש 6 + נת״ע', 'PFI',
                   'פעיל - ללא הסמכה מוגדרת', 'חלת - מחלה']
    # Other sheets for reference
    other_sheets = ['משימות להמשך טיפול', 'ללא הסמכה - לבירור',
                    'ריכוז כל המשימות', 'משימות לפי אחראי', 'סיכום כללי']

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = read_sheet_data(ws)
        header_idx, id_col_idx = find_header_and_id_col(rows)

        if header_idx is not None:
            employee_count, ids = extract_employees(rows, header_idx, id_col_idx)
            results[sheet_name] = {'count': employee_count, 'ids': ids, 'is_cert': sheet_name in cert_sheets}
            all_ids.update(ids)
        else:
            # Summary sheet - parse manually
            if sheet_name == 'סיכום כללי':
                summary = {}
                for row in rows:
                    if row[0] is not None and row[1] is not None:
                        try:
                            summary[str(row[0])] = int(row[1])
                        except (ValueError, TypeError):
                            pass
                results[sheet_name] = {'count': 0, 'ids': set(), 'is_cert': False, 'summary': summary}
            else:
                results[sheet_name] = {'count': 0, 'ids': set(), 'is_cert': False}

    wb.close()
    return results, all_ids

def build_comparison():
    pikoh_results, pikoh_ids = analyze_pikoh()
    ka_results, ka_ids = analyze_ka()

    output = []

    # --- PIKOH FILE ---
    output.append("# Pikoh File (Source of Truth): מאושרי נת\"ע לשיבוץ")
    output.append("")
    for sheet_name, data in pikoh_results.items():
        output.append(f"- Sheet '{sheet_name}': {data['count']} employees, {len(data['ids'])} unique IDs")
    output.append(f"- **Total unique IDs: {len(pikoh_ids)}**")
    output.append("")

    # --- KA FILE ---
    output.append("# KA File: כ\"א +משימות")
    output.append("")

    # Cert type sheets
    cert_type_ids = {}
    output.append("## Certification Type Sheets (primary employee lists)")
    for sheet_name in ['מאושרי נת״ע', 'מאושרי כביש 6', 'מאושרי כביש 6 + נת״ע', 'PFI',
                       'פעיל - ללא הסמכה מוגדרת', 'חלת - מחלה']:
        if sheet_name in ka_results:
            d = ka_results[sheet_name]
            output.append(f"- {sheet_name}: {d['count']} employees, {len(d['ids'])} unique IDs")
            cert_type_ids[sheet_name] = d['ids']

    # Count unique across cert sheets only
    all_cert_ids = set()
    for ids in cert_type_ids.values():
        all_cert_ids.update(ids)
    output.append(f"- **Total unique across cert sheets: {len(all_cert_ids)}**")
    output.append("")

    # Other sheets
    output.append("## Other Sheets (tasks, summaries)")
    for sheet_name in ['משימות להמשך טיפול', 'ללא הסמכה - לבירור', 'ריכוז כל המשימות', 'משימות לפי אחראי']:
        if sheet_name in ka_results:
            d = ka_results[sheet_name]
            output.append(f"- {sheet_name}: {d['count']} employees, {len(d['ids'])} unique IDs")
    output.append(f"- **Total unique across ALL KA sheets: {len(ka_ids)}**")
    output.append("")

    # --- OVERLAP ANALYSIS ---
    output.append("# Cross-File Overlap")
    common = pikoh_ids & ka_ids
    only_pikoh = pikoh_ids - ka_ids
    only_ka = ka_ids - pikoh_ids
    output.append(f"- Pikoh unique IDs: {len(pikoh_ids)}")
    output.append(f"- KA unique IDs (all sheets): {len(ka_ids)}")
    output.append(f"- Common (in both): {len(common)}")
    output.append(f"- Only in Pikoh: {len(only_pikoh)}")
    output.append(f"- Only in KA: {len(only_ka)}")
    output.append("")

    if only_pikoh:
        output.append(f"IDs only in Pikoh ({len(only_pikoh)}): {sorted(only_pikoh)}")
    if only_ka and len(only_ka) <= 50:
        output.append(f"IDs only in KA ({len(only_ka)}): {sorted(only_ka)}")
    output.append("")

    # --- Overlap between Pikoh and specific KA cert sheets ---
    output.append("# Pikoh vs KA Cert Sheet Overlap")
    for sheet_name, ids in cert_type_ids.items():
        overlap = pikoh_ids & ids
        output.append(f"- Pikoh ∩ '{sheet_name}': {len(overlap)} common IDs out of {len(ids)} in KA sheet")
    output.append("")

    # --- DB COMPARISON ---
    db_counts = {
        'חוצה ישראל': 0,
        'חוצה צפון (PFI)': 1,
        'כביש 6': 20,
        'כביש 6 + נת״ע': 0,
        'נת״ע': 76,
        'נתיבי ישראל': 0,
    }
    db_total_employees = 156
    db_unique_with_certs = 90

    output.append("# DB vs Excel Comparison")
    output.append("")
    output.append("| Source | נת״ע | כביש 6 | כביש 6 + נת״ע | PFI | ללא הסמכה | חלת/מחלה | Total Unique |")
    output.append("|--------|------|--------|---------------|-----|-----------|----------|-------------|")

    # Pikoh row
    pikoh_count = pikoh_results.get('מאושרי נת״ע לשיבוץ', {}).get('count', 0)
    pikoh_id_count = len(pikoh_ids)
    output.append(f"| Pikoh file | {pikoh_count} | - | - | - | - | - | {pikoh_id_count} |")

    # KA rows by cert type
    ka_natav = ka_results.get('מאושרי נת״ע', {})
    ka_kvish6 = ka_results.get('מאושרי כביש 6', {})
    ka_kvish6_natav = ka_results.get('מאושרי כביש 6 + נת״ע', {})
    ka_pfi = ka_results.get('PFI', {})
    ka_no_cert = ka_results.get('פעיל - ללא הסמכה מוגדרת', {})
    ka_sick = ka_results.get('חלת - מחלה', {})

    output.append(f"| KA file | {ka_natav.get('count', 0)} | {ka_kvish6.get('count', 0)} | {ka_kvish6_natav.get('count', 0)} | {ka_pfi.get('count', 0)} | {ka_no_cert.get('count', 0)} | {ka_sick.get('count', 0)} | {len(all_cert_ids)} |")

    # DB row - map cert types
    output.append(f"| DB (certs) | {db_counts['נת״ע']} | {db_counts['כביש 6']} | {db_counts['כביש 6 + נת״ע']} | {db_counts['חוצה צפון (PFI)']} | - | - | {db_unique_with_certs} |")
    output.append(f"| DB (total employees) | - | - | - | - | - | - | {db_total_employees} |")
    output.append("")

    # --- Detailed Delta Analysis ---
    output.append("# Delta Analysis: KA File vs DB")
    output.append("")
    output.append("| Cert Type | KA File | DB Certs | Delta | Notes |")
    output.append("|-----------|---------|----------|-------|-------|")

    ka_natav_count = ka_natav.get('count', 0)
    output.append(f"| נת״ע | {ka_natav_count} | {db_counts['נת״ע']} | {db_counts['נת״ע'] - ka_natav_count:+d} | DB has more certs than KA employees |")

    ka_k6_count = ka_kvish6.get('count', 0)
    output.append(f"| כביש 6 | {ka_k6_count} | {db_counts['כביש 6']} | {db_counts['כביש 6'] - ka_k6_count:+d} | DB has more certs (combo type employees?) |")

    ka_k6n_count = ka_kvish6_natav.get('count', 0)
    output.append(f"| כביש 6 + נת״ע | {ka_k6n_count} | {db_counts['כביש 6 + נת״ע']} | {db_counts['כביש 6 + נת״ע'] - ka_k6n_count:+d} | Legacy type in DB, KA has {ka_k6n_count} |")

    ka_pfi_count = ka_pfi.get('count', 0)
    output.append(f"| PFI (חוצה צפון) | {ka_pfi_count} | {db_counts['חוצה צפון (PFI)']} | {db_counts['חוצה צפון (PFI)'] - ka_pfi_count:+d} | Match |")

    output.append(f"| חוצה ישראל | - | {db_counts['חוצה ישראל']} | - | Not in KA file |")
    output.append(f"| נתיבי ישראל | - | {db_counts['נתיבי ישראל']} | - | Not in KA file |")
    output.append("")

    # Total employees
    output.append("# Employee Count Summary")
    output.append("")
    output.append(f"- **Pikoh file**: {pikoh_id_count} unique employees (all נת״ע)")
    output.append(f"- **KA file cert sheets**: {len(all_cert_ids)} unique employees across cert types")
    output.append(f"- **KA file all sheets**: {len(ka_ids)} unique IDs (includes tasks, no-cert, sick leave)")
    output.append(f"- **DB total employees**: {db_total_employees}")
    output.append(f"- **DB employees with certs**: {db_unique_with_certs}")
    output.append(f"- **DB total certifications**: {sum(db_counts.values())} (97)")
    output.append("")

    # Check: how many KA cert-sheet employees overlap with DB
    # We know DB has 76 נת״ע certs, KA has 56 נת״ע employees
    # The 76 in DB likely includes the 6 כביש 6 + נת״ע employees counted under נת״ע cert
    output.append("# Key Observations")
    output.append("")
    output.append("1. **DB נת״ע cert count (76) > KA נת״ע sheet (56)**: The DB likely counts each certification separately,")
    output.append("   including the 6 'כביש 6 + נת״ע' employees who also have a נת״ע cert. 56 + 6 = 62, still 14 short of 76.")
    output.append("   The remaining gap may be from employees in other sheets (tasks, no-cert) who have been assigned a נת״ע cert in the DB.")
    output.append("")
    output.append("2. **DB כביש 6 cert count (20) > KA כביש 6 sheet (14)**: Similarly, the 6 'כביש 6 + נת״ע' employees")
    output.append("   also have a כביש 6 cert. 14 + 6 = 20, which matches exactly!")
    output.append("")
    output.append("3. **DB כביש 6 + נת״ע (0 certs)**: This legacy combo type has 0 certs in DB, as expected -")
    output.append("   those employees' certs were split into separate נת״ע and כביש 6 certs.")
    output.append("")
    output.append("4. **DB total employees (156) vs KA all sheets (148)**: DB has 8 more employee records.")
    output.append("   These could be employees added directly to the DB that aren't in the KA file.")
    output.append("")
    output.append("5. **Pikoh file (71 employees)**: This is a subset of the נת״ע certified employees,")
    output.append("   specifically those approved for placement (שיבוץ). Not all נת״ע employees are in this list.")
    output.append("")

    return '\n'.join(output), pikoh_results, ka_results, pikoh_ids, ka_ids, all_cert_ids

if __name__ == '__main__':
    report, pikoh_results, ka_results, pikoh_ids, ka_ids, all_cert_ids = build_comparison()

    out_path = r"C:\Users\maor4\OneDrive\Desktop\certimanager\.claude\worktrees\affectionate-goldstine\scripts\analysis_final.txt"
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(report)

    print(f"Report written to {out_path}")

    # Also output JSON summary for programmatic use
    summary = {
        'pikoh': {
            'total_unique_ids': len(pikoh_ids),
            'sheets': {k: {'count': v['count'], 'unique_ids': len(v['ids'])} for k, v in pikoh_results.items()}
        },
        'ka': {
            'total_unique_ids': len(ka_ids),
            'cert_sheet_unique_ids': len(all_cert_ids),
            'sheets': {k: {'count': v['count'], 'unique_ids': len(v['ids'])} for k, v in ka_results.items()}
        }
    }

    json_path = r"C:\Users\maor4\OneDrive\Desktop\certimanager\.claude\worktrees\affectionate-goldstine\scripts\analysis_summary.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"JSON summary written to {json_path}")
